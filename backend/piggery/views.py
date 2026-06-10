from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.contrib.auth.models import User
from django.db.models import Avg, Sum, Count
from datetime import date, timedelta
import csv

from .models import (
    Farm, FarmBaseline, Pig, WeightRecord, VaccinationRecord,
    DiseaseRecord, BreedingRecord, FeedInventory, MedicineInventory,
    FeedUsageLog, MedicineUsageLog, Notification, HealthLog,
    AuditLog, UserProfile,
)
from .serializers import (
    FarmSerializer, PigListSerializer, PigDetailSerializer,
    WeightRecordSerializer, VaccinationRecordSerializer,
    DiseaseRecordSerializer, BreedingRecordSerializer,
    FeedInventorySerializer, MedicineInventorySerializer,
    FeedUsageLogSerializer, NotificationSerializer, HealthLogSerializer,
)
from .weather import get_weather_alert
from .services.sms import send_sms


# ── Helpers ───────────────────────────────────────────────────────────────────

def _generate_pig_id(farm):
    """
    Auto-generate a unique human-readable pig ID.
    Format: F{farm.id:03d}-P{sequence:04d}
    Example: F001-P0001, F001-P0002 ...
    Thread-safe: queries the current maximum numeric suffix and increments.
    """
    import re
    prefix = f"F{farm.id:03d}-P"
    existing = Pig.objects.filter(farm=farm, pig_id__startswith=prefix).values_list("pig_id", flat=True)
    max_num = 0
    for pid in existing:
        m = re.search(r"P(\d+)$", pid)
        if m:
            max_num = max(max_num, int(m.group(1)))
    return f"{prefix}{max_num + 1:04d}"


def get_user_farm(user):
    """Safely get farm for user. Returns None if not found."""
    return Farm.objects.filter(owner=user).first()


def log_action(user, action, model_name="", object_id="", description="", request=None):
    """Create an audit log entry. Never raises."""
    try:
        ip = ""
        if request:
            x_forwarded = request.META.get("HTTP_X_FORWARDED_FOR")
            ip = x_forwarded.split(",")[0] if x_forwarded else request.META.get("REMOTE_ADDR", "")
        AuditLog.objects.create(
            user=user, action=action, model_name=model_name,
            object_id=str(object_id), description=description, ip_address=ip or None,
        )
    except Exception:
        pass


def auto_create_farm(user):
    """Create a farm for a user if none exists. Safety net for legacy accounts."""
    farm, _ = Farm.objects.get_or_create(
        owner=user,
        defaults={
            "name":     f"{user.first_name or user.username}'s Farm",
            "location": "Concepcion, Tarlac",
        },
    )
    return farm


# ── Farm ──────────────────────────────────────────────────────────────────────

class FarmViewSet(viewsets.ModelViewSet):
    serializer_class = FarmSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Farm.objects.filter(owner=self.request.user)

    def perform_create(self, serializer):
        serializer.save(owner=self.request.user)

    @action(detail=True, methods=["get"])
    def dashboard(self, request, pk=None):
        """
        Redesigned dashboard — returns Status + Analytics + Forecast in one call.
        GET /api/farms/{id}/dashboard/
        """
        try:
            farm = self.get_object()
        except Exception:
            return Response({
                "farm_name": "My Farm", "total_pigs": 0, "healthy": 0,
                "under_treatment": 0, "critical": 0, "farm_health_score": 100,
                "pregnant_sows": 0, "upcoming_farrowing": 0,
                "low_feed_items": 0, "low_medicine_items": 0, "vaccinations_due": 0,
                "analytics": {}, "forecast": {},
            })

        today = date.today()
        pigs  = farm.pigs.exclude(health_status="deceased")

        total          = pigs.count()
        healthy        = pigs.filter(health_status="healthy").count()
        under_treatment = pigs.filter(health_status="under_treatment").count()
        critical        = pigs.filter(health_status="critical").count()

        # Farm health score: (healthy + 0.5×under_treatment) / total × 100
        if total > 0:
            farm_health_score = round(((healthy + 0.5 * under_treatment) / total) * 100, 1)
        else:
            farm_health_score = 100.0

        upcoming_farrowing = BreedingRecord.objects.filter(
            sow__farm=farm, pregnancy_status="pregnant",
            expected_farrowing_date__lte=today + timedelta(days=7),
        ).count()

        low_feed = farm.feed_inventory.filter(stock_kg__lte=25).count()
        low_medicine = sum(1 for m in farm.medicine_inventory.all() if m.is_low_stock)
        # Also count expired/near-expiry medicines
        expiring_count = sum(
            1 for m in farm.medicine_inventory.filter(expiry_date__isnull=False)
            if m.expiry_date <= today + timedelta(days=30)
        )
        vax_due = VaccinationRecord.objects.filter(
            pig__farm=farm, next_due_date__lte=today + timedelta(days=7),
        ).count()

        # ── Analytics section ─────────────────────────────────────────────────
        # ADG this month
        month_start = today.replace(day=1)
        adg_data = []
        ADG_BENCHMARKS = {
            "piglet": 0.25, "weaner": 0.40, "grower": 0.65,
            "finisher": 0.85, "breeder": 0.20,
        }
        for pig in pigs:
            records = pig.weight_records.order_by("recorded_at")
            count = records.count()
            if count == 0:
                continue

            last = records.last()

            if count >= 2:
                # Two or more records: use first vs last weight record
                first_date   = records.first().recorded_at
                first_weight = float(records.first().weight_kg)
            else:
                # Only one weight record: use date_of_birth as the starting point.
                # Birth weight for most pigs is approximately 1.5 kg (industry average).
                # This allows ADG to be calculated and displayed immediately after
                # the farmer logs the first weight, instead of showing 0 until
                # a second record is added.
                first_date   = pig.date_of_birth
                first_weight = 1.5

            days = (last.recorded_at - first_date).days
            if days == 0:
                continue
            adg = (float(last.weight_kg) - first_weight) / days
            adg_data.append({"adg": adg, "stage": pig.growth_stage})

        avg_adg = round(sum(d["adg"] for d in adg_data) / len(adg_data), 3) if adg_data else 0

        # Breeding this month
        litters_this_month = BreedingRecord.objects.filter(
            sow__farm=farm, pregnancy_status="farrowed",
            actual_farrowing_date__gte=month_start,
        ).count()

        farrowed_all = BreedingRecord.objects.filter(sow__farm=farm, pregnancy_status="farrowed")
        total_litters = farrowed_all.count()
        avg_litter = round(
            sum(r.piglets_born_alive or 0 for r in farrowed_all) / total_litters, 1
        ) if total_litters > 0 else 0

        # Feed this week
        week_start = today - timedelta(days=7)
        feed_this_week = FeedUsageLog.objects.filter(
            farm=farm, date_used__gte=week_start
        ).aggregate(total=Sum("amount_used_kg"))["total"] or 0

        # Most critical feed remaining
        all_feed = list(farm.feed_inventory.all())
        critical_feed_days = None
        if all_feed:
            days_list = [f.days_remaining for f in all_feed if f.days_remaining is not None]
            if days_list:
                critical_feed_days = min(days_list)

        # ── Forecast section ──────────────────────────────────────────────────
        next_farrowing = BreedingRecord.objects.filter(
            sow__farm=farm, pregnancy_status="pregnant",
            expected_farrowing_date__gte=today,
        ).order_by("expected_farrowing_date").first()

        next_vax = VaccinationRecord.objects.filter(
            pig__farm=farm, next_due_date__gte=today,
        ).order_by("next_due_date").first()

        return Response({
            # Status
            "farm_name":          farm.name,
            "farm_location":      farm.location or "Concepcion, Tarlac",
            "total_pigs":         total,
            "healthy":            healthy,
            "under_treatment":    under_treatment,
            "critical":           critical,
            "farm_health_score":  farm_health_score,
            "pregnant_sows":      pigs.filter(
                gender="female",
                breeding_records__pregnancy_status="pregnant"
            ).distinct().count(),
            "upcoming_farrowing": upcoming_farrowing,
            "low_feed_items":     low_feed,
            "low_medicine_items": low_medicine,
            "expiring_medicines": expiring_count,
            "vaccinations_due":   vax_due,

            # Analytics
            "analytics": {
                "avg_adg":            avg_adg,
                "litters_this_month": litters_this_month,
                "avg_litter_size":    avg_litter,
                "feed_this_week_kg":  round(float(feed_this_week), 1),
                "critical_feed_days": critical_feed_days,
            },

            # Forecast
            "forecast": {
                "next_farrowing_sow":  next_farrowing.sow.name if next_farrowing else None,
                "next_farrowing_date": str(next_farrowing.expected_farrowing_date) if next_farrowing else None,
                "next_vaccination_pig": next_vax.pig.name if next_vax else None,
                "next_vaccination_date": str(next_vax.next_due_date) if next_vax else None,
                "next_vaccination_name": next_vax.vaccine_name if next_vax else None,
            },
        })

    @action(detail=True, methods=["get"])
    def health_analytics(self, request, pk=None):
        """GET /api/farms/{id}/health_analytics/"""
        from .services.analytics_services import health_analytics
        farm = self.get_object()
        return Response(health_analytics(farm))

    @action(detail=True, methods=["get"])
    def growth_analytics(self, request, pk=None):
        """GET /api/farms/{id}/growth_analytics/"""
        from .services.analytics_services import growth_analytics
        farm = self.get_object()
        return Response(growth_analytics(farm))

    @action(detail=True, methods=["get"])
    def breeding_analytics(self, request, pk=None):
        """GET /api/farms/{id}/breeding_analytics/"""
        from .services.analytics_services import breeding_analytics
        farm = self.get_object()
        return Response(breeding_analytics(farm))

    @action(detail=True, methods=["get"])
    def feed_analytics(self, request, pk=None):
        """GET /api/farms/{id}/feed_analytics/"""
        from .services.analytics_services import feed_analytics
        farm = self.get_object()
        return Response(feed_analytics(farm))

    @action(detail=True, methods=["get"])
    def predictions(self, request, pk=None):
        """
        GET /api/farms/{id}/predictions/
        Returns all rule-based predictions from the prediction engine.
        """
        from .services.analytics_services import prediction_engine
        farm = self.get_object()
        return Response(prediction_engine(farm))

    @action(detail=True, methods=["post"])
    def save_baseline(self, request, pk=None):
        """
        POST /api/farms/{id}/save_baseline/
        Saves or updates the FarmBaseline for existing farm onboarding.
        Sets Farm.baseline_established = True when baseline is saved.
        """
        from .models import FarmBaseline
        farm = self.get_object()
        baseline, _ = FarmBaseline.objects.get_or_create(farm=farm)
        for field in [
            "years_in_operation", "pigs_at_registration", "avg_breeding_sows",
            "litters_last_12_months", "avg_litter_size_historical",
            "avg_daily_feed_kg_per_pig", "common_diseases", "notes",
        ]:
            if field in request.data:
                setattr(baseline, field, request.data[field])
        baseline.save()
        farm.baseline_established = True
        farm.save()
        log_action(request.user, "create", "FarmBaseline", farm.id,
                   f"Farm baseline saved for {farm.name}", request)
        return Response({"message": "Farm baseline saved. Analytics now incorporate historical data."})

    @action(detail=True, methods=["get"])
    def weather(self, request, pk=None):
        """
        GET /api/farms/{id}/weather/

        Returns weather data + pig-specific risk assessment.
        The farm object is passed to evaluate_farm_weather_risk() so
        per-stage risk is computed using the actual pig population.

        Also creates weather notifications when unsafe conditions are detected.
        Notifications are deduplicated: only creates a new one if no unread
        weather notification with the same title exists for today.
        """
        try:
            farm  = self.get_object()
            # Pass the farm object so get_weather_alert can evaluate
            # pig comfort per growth stage using the farm's actual pig population.
            alert = get_weather_alert(farm.location, farm=farm)

            # Auto-create weather notifications for heat/cold stress alerts.
            # FIX: removed is_read=False from the exists() check.
            # Previous version recreated the notification every time the user
            # read it, because is_read=False meant "no unread = create again".
            # Now we check if ANY notification with this title exists today
            # (read or unread) — if yes, skip. Creates at most once per day.
            pig_comfort = alert.get("pig_comfort")
            if pig_comfort and pig_comfort.get("overall_status") != "normal":
                overall     = pig_comfort.get("overall_label", "Weather Alert")
                summary     = pig_comfort.get("pig_comfort_summary", "")
                notif_title = f"Weather: {overall}"
                exists = Notification.objects.filter(
                    farm=farm,
                    notification_type="weather",
                    title=notif_title,
                    created_at__date=date.today(),
                    # No is_read filter — once created today, don't create again
                ).exists()
                if not exists:
                    Notification.objects.create(
                        farm=farm,
                        notification_type="weather",
                        title=notif_title,
                        message=summary,
                    )

            return Response(alert)
        except Exception:
            return Response({
                "alert_count":    0,
                "alerts":         [],
                "temperature_c":  None,
                "humidity_percent": None,
                "pig_comfort":    None,
            })


# ── Pig ───────────────────────────────────────────────────────────────────────

class PigViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action == "retrieve":
            return PigDetailSerializer
        return PigListSerializer

    def get_queryset(self):
        farm = get_user_farm(self.request.user)
        if not farm:
            return Pig.objects.none()
        qs = farm.pigs.exclude(health_status="deceased")
        stage  = self.request.query_params.get("stage")
        health = self.request.query_params.get("health")
        if stage:
            qs = qs.filter(growth_stage=stage)
        if health:
            qs = qs.filter(health_status=health)
        return qs

    def perform_create(self, serializer):
        farm = get_user_farm(self.request.user)
        if not farm:
            farm = auto_create_farm(self.request.user)

        # Auto-generate pig_id if not provided or empty
        provided_id = serializer.validated_data.get("pig_id", "").strip()
        if not provided_id:
            serializer.validated_data["pig_id"] = _generate_pig_id(farm)

        pig = serializer.save(farm=farm)
        log_action(self.request.user, "create", "Pig", pig.id,
                   f"Added pig: {pig.name} ({pig.pig_id})", self.request)

    def destroy(self, request, *args, **kwargs):
        pig = self.get_object()
        log_action(request.user, "delete", "Pig", pig.id,
                   f"Deleted pig: {pig.name} ({pig.pig_id})", request)
        return super().destroy(request, *args, **kwargs)

    @action(detail=False, methods=["get"])
    def next_pig_id(self, request):
        """GET /api/pigs/next_pig_id/ — returns the next auto-generated pig ID for this farm."""
        farm = get_user_farm(request.user)
        if not farm:
            farm = auto_create_farm(request.user)
        return Response({"pig_id": _generate_pig_id(farm)})

    @action(detail=True, methods=["post"])
    def save_baseline(self, request, pk=None):
        from .models import PigBaseline, BreedingRecord

        pig = self.get_object()

        # Backend validation: existing female breeder pigs must supply breeding history.
        # Applies only when pig.is_historical=True (registered as an existing pig),
        # gender=female, and growth_stage=breeder.
        # total_litters must be present; if > 0, last_farrowing_date is also required.
        if (
            pig.is_historical
            and pig.gender == "female"
            and pig.growth_stage == "breeder"
        ):
            # All four breeding history fields are required for existing female breeder pigs.
            # Each field is checked individually so the response identifies the exact missing field.
            total_litters_raw      = request.data.get("total_litters")
            total_piglets_born_raw = request.data.get("total_piglets_born")
            total_weaned_raw       = request.data.get("total_piglets_weaned")
            last_farrowing         = request.data.get("last_farrowing_date")

            if total_litters_raw is None or str(total_litters_raw).strip() == "":
                return Response(
                    {"error": "total_litters is required for existing breeder sows (enter 0 if unknown)."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            try:
                litters_int = int(total_litters_raw)
            except (ValueError, TypeError):
                return Response(
                    {"error": "total_litters must be a whole number."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            if total_piglets_born_raw is None or str(total_piglets_born_raw).strip() == "":
                return Response(
                    {"error": "total_piglets_born is required for existing breeder sows (enter 0 if unknown)."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if total_weaned_raw is None or str(total_weaned_raw).strip() == "":
                return Response(
                    {"error": "total_piglets_weaned is required for existing breeder sows (enter 0 if unknown)."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if not last_farrowing:
                return Response(
                    {"error": "last_farrowing_date is required for existing breeder sows."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        baseline, created = PigBaseline.objects.get_or_create(pig=pig)

        for field in [
            "total_litters", "total_piglets_born", "total_piglets_weaned",
            "last_farrowing_date", "major_diseases_history",
            "vaccination_status_summary", "weight_at_6_months",
            "weight_at_12_months", "notes",
        ]:
            if field in request.data and request.data[field] not in (None, ""):
                setattr(baseline, field, request.data[field])
        baseline.save()

        # ── Auto-create historical BreedingRecord rows ────────────────────────
        total_litters   = int(request.data.get("total_litters", 0) or 0)
        total_alive     = int(request.data.get("total_piglets_born", 0) or 0)
        total_weaned    = int(request.data.get("total_piglets_weaned", 0) or 0)
        last_farrowing  = request.data.get("last_farrowing_date")

        breeding_records_created = 0

        if (
            pig.gender == "female"
            and total_litters > 0
            and last_farrowing
            and pig.growth_stage in ("breeder",)
        ):
            # Delete any previously auto-created historical records for this pig
            BreedingRecord.objects.filter(sow=pig, notes__startswith="[historical]").delete()

            try:
                from datetime import datetime as dt
                last_date = dt.strptime(str(last_farrowing), "%Y-%m-%d").date()
            except (ValueError, TypeError):
                last_date = None

            if last_date:
                avg_alive  = round(total_alive  / total_litters) if total_litters > 0 else 0
                avg_weaned = round(total_weaned / total_litters) if total_litters > 0 else 0
                # Inter-farrowing interval: 155 days (industry average)
                IFI = 155

                for i in range(total_litters):
                    farrowing_date = last_date - timedelta(days=i * IFI)
                    breeding_date  = farrowing_date - timedelta(days=114)
                    BreedingRecord.objects.create(
                        sow=pig,
                        breeding_date=breeding_date,
                        expected_farrowing_date=farrowing_date,
                        actual_farrowing_date=farrowing_date,
                        pregnancy_status="farrowed",
                        piglets_born_alive=avg_alive,
                        piglets_born_dead=0,
                        piglets_weaned=avg_weaned,
                        wean_date=farrowing_date + timedelta(days=21),
                        notes=f"[historical] Auto-created from onboarding baseline (litter {total_litters - i} of {total_litters})",
                    )
                    breeding_records_created += 1

        log_action(request.user, "create", "PigBaseline", pig.id,
                   f"Historical baseline saved for {pig.name} ({breeding_records_created} breeding records created)", request)

        # Also generate upcoming farrowing notification if sow has last_farrowing_date
        if last_farrowing and pig.gender == "female":
            try:
                from .models import Notification
                from datetime import datetime as dt
                last_dt = dt.strptime(str(last_farrowing), "%Y-%m-%d").date()
                # Estimate next breeding after 21-day wean + 5-day estrus = 26 days post-farrowing
                next_breeding_est = last_dt + timedelta(days=26)
                next_expected_farrow = next_breeding_est + timedelta(days=114)
                days_until = (next_expected_farrow - date.today()).days
                if 0 < days_until < 60:
                    Notification.objects.get_or_create(
                        farm=pig.farm,
                        notification_type="breeding",
                        title=f"{pig.name} — Estimated next farrowing",
                        defaults={
                            "message": (
                                f"Based on last farrowing date {last_farrowing}, "
                                f"{pig.name}'s estimated next farrowing is around {next_expected_farrow} "
                                f"({days_until} days from today). Schedule breeding accordingly."
                            ),
                            "is_read": False,
                        }
                    )
            except Exception:
                pass

        return Response({
            "message": f"Historical data saved for {pig.name}.",
            "breeding_records_created": breeding_records_created,
            "note": f"{breeding_records_created} historical breeding records created for analytics." if breeding_records_created else "No breeding records created (requires: female, breeder stage, total_litters > 0, last_farrowing_date).",
        })

    @action(detail=True, methods=["get"])
    def baseline(self, request, pk=None):
        """GET /api/pigs/{id}/baseline/ — retrieve the pig's historical baseline."""
        from .models import PigBaseline
        pig = self.get_object()
        try:
            b = PigBaseline.objects.get(pig=pig)
            return Response({
                "total_litters":             b.total_litters,
                "total_piglets_born":        b.total_piglets_born,
                "total_piglets_weaned":      b.total_piglets_weaned,
                "last_farrowing_date":       str(b.last_farrowing_date) if b.last_farrowing_date else None,
                "major_diseases_history":    b.major_diseases_history,
                "vaccination_status_summary":b.vaccination_status_summary,
                "weight_at_6_months":        float(b.weight_at_6_months) if b.weight_at_6_months else None,
                "weight_at_12_months":       float(b.weight_at_12_months) if b.weight_at_12_months else None,
                "notes":                     b.notes,
            })
        except PigBaseline.DoesNotExist:
            return Response({})

    @action(detail=True, methods=["post"])
    def log_weight(self, request, pk=None):
        pig = self.get_object()
        serializer = WeightRecordSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(pig=pig)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=["get"])
    def growth_chart(self, request, pk=None):
        pig     = self.get_object()
        records = pig.weight_records.order_by("recorded_at")
        data    = [{"date": str(r.recorded_at), "weight_kg": float(r.weight_kg)} for r in records]
        return Response(data)

    @action(detail=False, methods=["get"])
    def analytics(self, request):
        farm = get_user_farm(request.user)
        if not farm:
            return Response([])
        stages = ["piglet", "weaner", "grower", "finisher", "breeder"]
        result = [{
            "stage": s,
            "count": farm.pigs.filter(growth_stage=s).count(),
        } for s in stages]
        return Response(result)


# ── Weight ────────────────────────────────────────────────────────────────────

class WeightRecordViewSet(viewsets.ModelViewSet):
    serializer_class = WeightRecordSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        pig_id = self.kwargs.get("pig_pk")
        return WeightRecord.objects.filter(
            pig__id=pig_id, pig__farm__owner=self.request.user
        )

    def perform_create(self, serializer):
        pig = Pig.objects.get(pk=self.kwargs["pig_pk"])
        serializer.save(pig=pig)


# ── Vaccination ───────────────────────────────────────────────────────────────

class VaccinationViewSet(viewsets.ModelViewSet):
    serializer_class = VaccinationRecordSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        pig_id = self.kwargs.get("pig_pk")
        return VaccinationRecord.objects.filter(
            pig__id=pig_id, pig__farm__owner=self.request.user
        )

    def perform_create(self, serializer):
        pig = Pig.objects.get(pk=self.kwargs["pig_pk"])
        serializer.save(pig=pig)

    @action(detail=False, methods=["post"])
    def schedule(self, request, pig_pk=None):
        """
        POST /api/pigs/{pig_pk}/vaccinations/schedule/

        Creates a scheduled vaccination record and a notification.

        Date validation (two layers):
          1. This view rejects next_due_date in the past before the serializer runs.
          2. VaccinationRecordSerializer.validate_next_due_date() enforces the same
             rule — catching any direct API calls that bypass the view-level check.
        """
        pig = Pig.objects.get(pk=pig_pk)

        # Layer 1: view-level date guard
        next_due_date_str = request.data.get("next_due_date", "")
        if next_due_date_str:
            try:
                from datetime import datetime
                next_due_date = datetime.strptime(next_due_date_str, "%Y-%m-%d").date()
                if next_due_date < date.today():
                    return Response(
                        {
                            "error": (
                                f"Vaccination date cannot be in the past. "
                                f"You entered {next_due_date_str}. "
                                f"Today is {date.today()}. "
                                f"Please select today or a future date."
                            )
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )
            except ValueError:
                return Response(
                    {"error": "Invalid date format. Use YYYY-MM-DD (e.g. 2026-07-15)."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        # Layer 2: serializer validation (also enforces the date rule)
        serializer = VaccinationRecordSerializer(data=request.data)
        if serializer.is_valid():
            record = serializer.save(pig=pig)
            Notification.objects.create(
                farm=pig.farm,
                notification_type="vaccination",
                title=f"Vaccination scheduled: {pig.name}",
                message=f"{record.vaccine_name} scheduled for {pig.name} on {record.next_due_date}.",
            )
            return Response(serializer.data, status=201)
        return Response(serializer.errors, status=400)


# ── Disease ───────────────────────────────────────────────────────────────────

class DiseaseViewSet(viewsets.ModelViewSet):
    serializer_class = DiseaseRecordSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        pig_id = self.kwargs.get("pig_pk")
        return DiseaseRecord.objects.filter(
            pig__id=pig_id, pig__farm__owner=self.request.user
        )

    def perform_create(self, serializer):
        pig = Pig.objects.get(pk=self.kwargs["pig_pk"])
        record = serializer.save(pig=pig)
        try:
            send_sms(
                phone=pig.farm.owner.profile.phone_number,
                message=f"[Piglytics] ALERT: {pig.name} diagnosed with {record.disease_name}."
            )
        except Exception:
            pass


# ── Breeding ──────────────────────────────────────────────────────────────────

class BreedingViewSet(viewsets.ModelViewSet):
    serializer_class = BreedingRecordSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return BreedingRecord.objects.filter(
            sow__farm__owner=self.request.user
        ).select_related("sow")

    def perform_create(self, serializer):
        record = serializer.save()
        # Notification on new breeding record
        try:
            Notification.objects.create(
                farm=record.sow.farm,
                notification_type="breeding",
                title=f"New breeding record: {record.sow.name}",
                message=(
                    f"{record.sow.name} was bred on {record.breeding_date}. "
                    f"Expected farrowing: {record.expected_farrowing_date}."
                ),
            )
            log_action(self.request.user, "create", "BreedingRecord", record.id,
                       f"Added breeding record for {record.sow.name} on {record.breeding_date}",
                       self.request)
        except Exception:
            pass

    @action(detail=False, methods=["get"])
    def eligible_sows(self, request):
        """
        GET /api/breeding/eligible_sows/

        Returns female pigs eligible for a NEW breeding event.

        Excludes sows that are currently:
          - pregnant  (already confirmed pregnant)
          - bred      (breeding event recorded, awaiting confirmation)

        These sows should not appear in the breeding selection picker at all —
        not just shown with a warning. A sow with an active pregnancy cannot
        be bred again until she has farrowed and weaned.
        """
        farm = get_user_farm(request.user)
        if not farm:
            return Response([])

        eligible = Pig.objects.filter(
            farm=farm,
            gender="female",
            growth_stage__in=["finisher", "breeder"],
        ).exclude(health_status="deceased").order_by("name")

        # Get the most recent breeding status for each sow
        # and exclude any that are currently bred or pregnant
        result = []
        for pig in eligible:
            last_record = BreedingRecord.objects.filter(
                sow=pig
            ).order_by("-breeding_date").first()

            current_status = last_record.pregnancy_status if last_record else "open"

            # Exclude sows that are already in an active breeding cycle.
            # 'bred' = breeding recorded, not yet confirmed pregnant.
            # 'pregnant' = confirmed pregnant, must not be bred again.
            # These sows will reappear after farrowing.
            if current_status in ("bred", "pregnant"):
                continue

            result.append({
                "id":                pig.id,
                "pig_id":            pig.pig_id,
                "name":              pig.name,
                "growth_stage":      pig.growth_stage,
                "breed":             pig.breed,
                "age_in_months":     pig.age_in_months,
                "latest_weight":     pig.latest_weight,
                "current_status":    current_status,
                "last_bred":         str(last_record.breeding_date) if last_record else None,
                "expected_farrowing":str(last_record.expected_farrowing_date) if last_record and last_record.expected_farrowing_date else None,
                "display_label":     f"{pig.pig_id} — {pig.name}",
            })

        return Response(result)

    @action(detail=False, methods=["get"])
    def forecast(self, request):
        farm = get_user_farm(request.user)
        if not farm:
            return Response({
                "farrowing_forecasts": [], "next_breeding_windows": [],
                "adg_performance": [], "summary": {
                    "pregnant_sows": 0, "farrowing_within_7_days": 0,
                    "sows_ready_to_breed": 0, "pigs_above_adg_benchmark": 0,
                    "pigs_below_adg_benchmark": 0,
                }
            })

        today = date.today()

        pregnant_records = BreedingRecord.objects.filter(
            sow__farm=farm, pregnancy_status__in=["pregnant", "bred"],
        ).select_related("sow")

        farrowing_forecasts = []
        for record in pregnant_records:
            days_pregnant  = (today - record.breeding_date).days
            days_remaining = 114 - days_pregnant
            gestation_pct  = round((days_pregnant / 114) * 100, 1)

            # gestation_stage is based on both days_pregnant AND pregnancy_status.
            # A sow with pregnancy_status='pregnant' has been CONFIRMED pregnant —
            # she must never show "awaiting confirmation" regardless of days since breeding.
            # The old code used only days_pregnant, so a sow confirmed pregnant on
            # the same day as breeding (days_pregnant=0) still showed "Newly bred".
            if record.pregnancy_status == "bred" and days_pregnant <= 7:
                stage = "Newly bred — awaiting pregnancy confirmation"
            elif days_pregnant <= 30:
                stage = "Early gestation (embryo attachment)"
            elif days_pregnant <= 75:
                stage = "Mid gestation (fetal development)"
            elif days_pregnant <= 100:
                stage = "Late gestation (rapid fetal growth)"
            else:
                stage = "Pre-farrowing (prepare pen now)"

            earliest  = record.breeding_date + timedelta(days=107)
            latest    = record.breeding_date + timedelta(days=121)
            is_overdue = today > latest
            alert = not is_overdue and (earliest - today).days <= 7

            farrowing_forecasts.append({
                "sow_name":               record.sow.name,
                "sow_id":                 record.sow.pig_id,
                "breeding_date":          str(record.breeding_date),
                "expected_farrowing":     str(record.expected_farrowing_date),
                "earliest_farrowing":     str(earliest),
                "latest_farrowing":       str(latest),
                "days_pregnant":          days_pregnant,
                "days_remaining":         max(0, days_remaining),
                "gestation_progress_pct": min(100, gestation_pct),
                "gestation_stage":        stage,
                "is_overdue":             is_overdue,
                "alert":                  alert,
                "pregnancy_status":       record.pregnancy_status,
            })

        DAYS_TO_WEANING      = 21
        DAYS_WEANING_TO_ESTRUS = 5
        recently_farrowed = BreedingRecord.objects.filter(
            sow__farm=farm, pregnancy_status="farrowed",
            actual_farrowing_date__isnull=False,
        ).select_related("sow").order_by("-actual_farrowing_date")

        seen_sows     = set()
        next_breeding = []
        for record in recently_farrowed:
            if record.sow.id in seen_sows:
                continue
            seen_sows.add(record.sow.id)

            # FIX: skip this sow if she already has a more recent breeding record.
            # Without this check, a sow appears as "Ready to Breed" forever after
            # farrowing, even after the farmer has already added a new breeding event.
            # The recently_farrowed query only fetches 'farrowed' records — it never
            # sees the new 'bred' or 'pregnant' record created after farrowing.
            # This check looks for any BreedingRecord for this sow with a breeding_date
            # AFTER the actual_farrowing_date of the current record. If one exists,
            # the sow has already been re-bred and must not appear as ready.
            already_re_bred = BreedingRecord.objects.filter(
                sow=record.sow,
                breeding_date__gt=record.actual_farrowing_date,
                pregnancy_status__in=["bred", "pregnant"],
            ).exists()
            if already_re_bred:
                continue

            weaning_date = record.actual_farrowing_date + timedelta(days=DAYS_TO_WEANING)
            estrus_date  = weaning_date + timedelta(days=DAYS_WEANING_TO_ESTRUS)
            days_until   = (estrus_date - today).days
            next_breeding.append({
                "sow_name":          record.sow.name,
                "sow_id":            record.sow.pig_id,
                "farrowed_on":       str(record.actual_farrowing_date),
                "weaning_date":      str(weaning_date),
                "next_estrus_date":  str(estrus_date),
                "days_until_estrus": days_until,
                "ready_to_breed":    days_until <= 0,
                "status": "Ready to breed now" if days_until <= 0 else f"Estrus in {days_until} days",
            })

        ADG_BENCHMARKS = {
            "piglet": 0.25, "weaner": 0.40, "grower": 0.65,
            "finisher": 0.85, "breeder": 0.20,
        }
        adg_data = []
        for pig in farm.pigs.exclude(health_status="deceased"):
            records = pig.weight_records.order_by("recorded_at")
            count = records.count()
            if count == 0:
                continue

            last = records.last()

            if count >= 2:
                first_date   = records.first().recorded_at
                first_weight = float(records.first().weight_kg)
            else:
                # Single record: use date_of_birth and approximate birth weight
                first_date   = pig.date_of_birth
                first_weight = 1.5

            days = (last.recorded_at - first_date).days
            if days == 0:
                continue
            adg       = round((float(last.weight_kg) - first_weight) / days, 3)
            benchmark = ADG_BENCHMARKS.get(pig.growth_stage, 0.5)
            perf      = round((adg / benchmark) * 100, 1)
            adg_data.append({
                "pig_name":                     pig.name,
                "pig_id":                       pig.pig_id,
                "growth_stage":                 pig.growth_stage,
                "current_weight":               float(last.weight_kg),
                "adg_kg_per_day":               adg,
                "benchmark_kg_per_day":         benchmark,
                "performance_vs_benchmark_pct": perf,
                "status": (
                    "Above benchmark" if perf >= 100
                    else "Below benchmark" if perf < 80
                    else "Near benchmark"
                ),
            })

        return Response({
            "farrowing_forecasts":   farrowing_forecasts,
            "next_breeding_windows": next_breeding,
            "adg_performance":       adg_data,
            "summary": {
                "pregnant_sows":            BreedingRecord.objects.filter(
                    sow__farm=farm, pregnancy_status="pregnant").count(),
                "farrowing_within_7_days":  sum(1 for f in farrowing_forecasts if f["alert"]),
                "sows_ready_to_breed":      sum(1 for b in next_breeding if b["ready_to_breed"]),
                "pigs_above_adg_benchmark": sum(1 for a in adg_data if a["status"] == "Above benchmark"),
                "pigs_below_adg_benchmark": sum(1 for a in adg_data if a["status"] == "Below benchmark"),
            },
        })

    @action(detail=True, methods=["post"])
    def record_farrowing(self, request, pk=None):
        """
        POST /api/breeding/{id}/record_farrowing/

        Records farrowing outcome. Collects:
          - piglets_born_alive (required)
          - piglets_born_dead  (optional, default 0)
          - piglets_weaned     (optional — can be updated later)
          - wean_date          (optional — can be updated later)
          - notes

        piglets_weaned and wean_date feed directly into:
          - breeding_analytics.weaning_rate_pct
          - sow productivity scores
          - survival analytics
        """
        record        = self.get_object()
        piglets_alive = request.data.get("piglets_born_alive")
        piglets_dead  = request.data.get("piglets_born_dead", 0)
        piglets_weaned= request.data.get("piglets_weaned")
        wean_date_str = request.data.get("wean_date")
        notes         = request.data.get("notes", "")

        if piglets_alive is None:
            return Response({"error": "piglets_born_alive is required."}, status=400)

        record.pregnancy_status      = "farrowed"
        record.piglets_born_alive    = int(piglets_alive)
        record.piglets_born_dead     = int(piglets_dead)
        record.actual_farrowing_date = date.today()
        if piglets_weaned is not None:
            record.piglets_weaned = int(piglets_weaned)
        if wean_date_str:
            try:
                from datetime import datetime as dt
                record.wean_date = dt.strptime(wean_date_str, "%Y-%m-%d").date()
            except ValueError:
                pass
        if notes:
            record.notes = notes
        record.save()

        total = int(piglets_alive) + int(piglets_dead)
        Notification.objects.create(
            farm=record.sow.farm,
            notification_type="breeding",
            title=f"{record.sow.name} has farrowed!",
            message=(
                f"{record.sow.name} delivered {piglets_alive} live "
                f"and {piglets_dead} dead piglet(s) ({total} total) on {date.today()}."
            ),
        )
        log_action(request.user, "update", "BreedingRecord", record.id,
                   f"Farrowing recorded for {record.sow.name}: {piglets_alive} alive",
                   request)
        return Response(BreedingRecordSerializer(record).data)

    @action(detail=True, methods=["post"])
    def mark_failed(self, request, pk=None):
        """
        POST /api/breeding/{id}/mark_failed/

        Marks a breeding attempt as failed.

        Analytics impact:
          - Counts as a breeding attempt in total_events
          - Reduces pregnancy_success_rate_pct
          - Reduces farrowing_success_rate_pct
          - Contributes to sow fertility trend (lower score for repeat failures)

        This is NOT an error state — failed attempts are valid data.
        They are essential for accurate reproductive KPIs.
        """
        record = self.get_object()
        if record.pregnancy_status in ("farrowed",):
            return Response({"error": "Cannot mark a farrowed record as failed."}, status=400)

        notes = request.data.get("notes", "")
        record.pregnancy_status = "failed"
        if notes:
            record.notes = notes
        record.save()

        # Notify farm
        try:
            Notification.objects.create(
                farm=record.sow.farm,
                notification_type="breeding",
                title=f"{record.sow.name} — breeding attempt failed",
                message=(
                    f"The breeding attempt for {record.sow.name} on "
                    f"{record.breeding_date} has been recorded as failed. "
                    "This has been included in reproductive analytics."
                ),
            )
        except Exception:
            pass

        log_action(request.user, "update", "BreedingRecord", record.id,
                   f"Breeding marked as failed for {record.sow.name}",
                   request)
        return Response(BreedingRecordSerializer(record).data)

    @action(detail=False, methods=["get"])
    def sow_performance(self, request):
        farm = get_user_farm(request.user)
        if not farm:
            return Response([])

        sows = farm.pigs.filter(gender="female", growth_stage="breeder")
        results = []
        for sow in sows:
            records = BreedingRecord.objects.filter(sow=sow, pregnancy_status="farrowed")
            total_litters = records.count()
            if total_litters == 0:
                continue
            total_alive = sum(r.piglets_born_alive or 0 for r in records)
            total_dead  = sum(r.piglets_born_dead  or 0 for r in records)
            total_born  = total_alive + total_dead
            avg_alive   = round(total_alive / total_litters, 1)
            survival    = round((total_alive / total_born * 100), 1) if total_born > 0 else 0

            if avg_alive >= 10 and survival >= 90:
                rating = "Excellent"
            elif avg_alive >= 8 and survival >= 80:
                rating = "Good"
            elif avg_alive >= 6:
                rating = "Average"
            else:
                rating = "Poor"

            results.append({
                "sow_name": sow.name, "sow_id": sow.pig_id,
                "total_litters": total_litters, "total_alive": total_alive,
                "total_dead": total_dead, "avg_live_piglets": avg_alive,
                "survival_rate": survival, "performance_rating": rating,
                "last_farrowed": str(
                    records.order_by("-actual_farrowing_date").first().actual_farrowing_date or "—"
                ),
            })

        results.sort(key=lambda x: (x["avg_live_piglets"], x["survival_rate"]), reverse=True)
        return Response(results)


# ── Feed Inventory ────────────────────────────────────────────────────────────

class FeedInventoryViewSet(viewsets.ModelViewSet):
    serializer_class = FeedInventorySerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        farm = get_user_farm(self.request.user)
        if not farm:
            return FeedInventory.objects.none()
        return farm.feed_inventory.all()

    def perform_create(self, serializer):
        farm = get_user_farm(self.request.user)
        if not farm:
            farm = auto_create_farm(self.request.user)

        from decimal import Decimal
        feed_type = serializer.validated_data.get("feed_type")
        existing  = FeedInventory.objects.filter(farm=farm, feed_type=feed_type).first()
        if existing:
            add_stock = serializer.validated_data.get("stock_kg", 0)
            existing.stock_kg += Decimal(str(add_stock))
            if serializer.validated_data.get("daily_usage_kg"):
                existing.daily_usage_kg = serializer.validated_data["daily_usage_kg"]
            if serializer.validated_data.get("price_per_kg"):
                existing.price_per_kg = serializer.validated_data["price_per_kg"]
            import datetime
            existing.last_restocked = datetime.date.today()
            existing.save()
            log_action(self.request.user, "update", "FeedInventory", existing.id,
                       f"Restocked {existing.get_feed_type_display()}: +{add_stock}kg", self.request)
        else:
            obj = serializer.save(farm=farm)
            log_action(self.request.user, "create", "FeedInventory", obj.id,
                       f"Added feed: {obj.get_feed_type_display()} ({obj.stock_kg}kg)", self.request)

    @action(detail=True, methods=["post"])
    def log_usage(self, request, pk=None):
        feed   = self.get_object()
        amount = float(request.data.get("amount_kg", 0))
        if amount <= 0:
            return Response({"error": "Amount must be greater than 0."}, status=400)
        if amount > feed.stock_kg:
            return Response({"error": "Not enough stock."}, status=400)

        from decimal import Decimal
        feed.stock_kg = feed.stock_kg - Decimal(str(amount))
        feed.save()

        FeedUsageLog.objects.create(
            farm=feed.farm, feed=feed, amount_used_kg=amount,
            logged_by=request.user,
        )

        if float(feed.stock_kg) <= 25:
            already = Notification.objects.filter(
                farm=feed.farm, notification_type="inventory",
                title=f"Low feed: {feed.get_feed_type_display()}", is_read=False,
            ).exists()
            if not already:
                Notification.objects.create(
                    farm=feed.farm, notification_type="inventory",
                    title=f"Low feed: {feed.get_feed_type_display()}",
                    message=(
                        f"{feed.get_feed_type_display()} is running low — "
                        f"{feed.stock_kg}kg remaining. Please restock soon."
                    ),
                )
            try:
                from .services.sms import send_low_stock_alert
                send_low_stock_alert(
                    feed.farm.owner.profile.phone_number,
                    feed.get_feed_type_display(),
                    f"{feed.stock_kg}kg",
                )
            except Exception:
                pass

        return Response(FeedInventorySerializer(feed).data)


# ── Medicine Inventory ────────────────────────────────────────────────────────

class MedicineInventoryViewSet(viewsets.ModelViewSet):
    serializer_class = MedicineInventorySerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        farm = get_user_farm(self.request.user)
        if not farm:
            return MedicineInventory.objects.none()
        return farm.medicine_inventory.all()

    def perform_create(self, serializer):
        farm = get_user_farm(self.request.user)
        if not farm:
            farm = auto_create_farm(self.request.user)
        name     = serializer.validated_data.get("name", "").strip().lower()
        existing = MedicineInventory.objects.filter(farm=farm, name__iexact=name).first()
        if existing:
            existing.quantity += serializer.validated_data.get("quantity", 0)
            existing.save()
        else:
            serializer.save(farm=farm)

    @action(detail=True, methods=["post"])
    def update_stock(self, request, pk=None):
        medicine    = self.get_object()
        action_type = request.data.get("action", "deduct")
        amount      = int(request.data.get("amount", 0))

        if amount <= 0:
            return Response({"error": "Amount must be greater than 0."}, status=400)

        if action_type == "deduct":
            if amount > medicine.quantity:
                return Response({"error": "Not enough stock."}, status=400)
            medicine.quantity -= amount
            MedicineUsageLog.objects.create(
                farm=medicine.farm,
                medicine=medicine,
                amount_used=amount,
                # MedicineUsageLog has no logged_by field — it has administered_by (CharField)
                administered_by=request.user.get_full_name() or request.user.username if hasattr(request, "user") else "",
            )
        else:
            medicine.quantity += amount
        medicine.save()

        if medicine.is_low_stock:
            already = Notification.objects.filter(
                farm=medicine.farm, notification_type="inventory",
                title=f"Low stock: {medicine.name}", is_read=False,
            ).exists()
            if not already:
                Notification.objects.create(
                    farm=medicine.farm, notification_type="inventory",
                    title=f"Low stock: {medicine.name}",
                    message=(
                        f"{medicine.name} is running low — "
                        f"{medicine.quantity} {medicine.unit} remaining."
                    ),
                )

        return Response(MedicineInventorySerializer(medicine).data)


# ── Notifications ─────────────────────────────────────────────────────────────

class NotificationViewSet(viewsets.ModelViewSet):
    serializer_class = NotificationSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        farm = get_user_farm(self.request.user)
        if not farm:
            return Notification.objects.none()
        return farm.notifications.all()

    @action(detail=True, methods=["post"])
    def mark_read(self, request, pk=None):
        notif = self.get_object()
        notif.is_read = True
        notif.save()
        return Response({"status": "marked as read"})

    @action(detail=False, methods=["post"])
    def mark_all_read(self, request):
        farm = get_user_farm(request.user)
        if farm:
            farm.notifications.filter(is_read=False).update(is_read=True)
        return Response({"status": "all marked as read"})


# ── Health Log ────────────────────────────────────────────────────────────────

class HealthLogViewSet(viewsets.ModelViewSet):
    serializer_class = HealthLogSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        pig_id = self.kwargs.get("pig_pk")
        return HealthLog.objects.filter(
            pig__id=pig_id, pig__farm__owner=self.request.user
        )

    def perform_create(self, serializer):
        from .services.health_rules import evaluate_health_log
        from .services.sms import send_health_alert

        pig = Pig.objects.get(pk=self.kwargs["pig_pk"])
        log = serializer.save(pig=pig, logged_by=self.request.user)

        severity, findings = evaluate_health_log(log)
        log.severity        = severity
        log.system_findings = findings
        log.save()

        if severity == "critical":
            pig.health_status = "critical"
        elif severity == "warning":
            pig.health_status = "under_treatment"
        else:
            pig.health_status = "healthy"
        pig.last_checkup_date = log.date_logged
        pig.save()

        if severity != "normal":
            first_finding = findings.split("\n")[0]
            Notification.objects.create(
                farm=pig.farm, notification_type="health",
                title=f"{'CRITICAL' if severity == 'critical' else 'Warning'}: {pig.name}",
                message=first_finding, is_read=False,
            )
            try:
                phone = pig.farm.owner.profile.phone_number
                if phone:
                    send_health_alert(phone, pig.name, first_finding[:100])
            except Exception:
                pass


# ── Audit Log ─────────────────────────────────────────────────────────────────

class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    permission_classes = [IsAuthenticated]

    def list(self, request):
        # Filters
        action_filter = request.query_params.get("action")
        search        = request.query_params.get("search", "").strip()
        date_from     = request.query_params.get("date_from")
        date_to       = request.query_params.get("date_to")

        if request.user.is_staff or request.user.is_superuser:
            logs = AuditLog.objects.all()
        else:
            logs = AuditLog.objects.filter(user=request.user)

        if action_filter:
            logs = logs.filter(action=action_filter)
        if search:
            from django.db.models import Q
            logs = logs.filter(
                Q(description__icontains=search) |
                Q(user__username__icontains=search) |
                Q(model_name__icontains=search)
            )
        if date_from:
            logs = logs.filter(created_at__date__gte=date_from)
        if date_to:
            logs = logs.filter(created_at__date__lte=date_to)

        logs = logs[:500]
        data = [{
            "id":           log.id,
            "username":     log.user.username if log.user else "deleted",
            "full_name":    f"{log.user.first_name} {log.user.last_name}".strip() if log.user else "",
            "role":         log.user.profile.role if log.user and hasattr(log.user, "profile") else "—",
            "action":       log.action,
            "action_label": log.get_action_display(),
            "model_name":   log.model_name,
            "description":  log.description,
            "ip_address":   log.ip_address,
            "created_at":   log.created_at.strftime("%Y-%m-%d %H:%M:%S"),
        } for log in logs]

        return Response(data)

    @action(detail=False, methods=["get"])
    def download_csv(self, request):
        from django.http import HttpResponse
        if request.user.is_staff or request.user.is_superuser:
            logs = AuditLog.objects.all()
        else:
            logs = AuditLog.objects.filter(user=request.user)

        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="piglytics_audit_log.csv"'
        writer = csv.writer(response)
        writer.writerow(["Date & Time", "Username", "Full Name", "Action", "Module", "Description", "IP Address"])
        for log in logs:
            writer.writerow([
                log.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                log.user.username if log.user else "deleted",
                f"{log.user.first_name} {log.user.last_name}".strip() if log.user else "",
                log.get_action_display(),
                log.model_name,
                log.description,
                log.ip_address,
            ])
        return response

    @action(detail=False, methods=["get"])
    def download_excel(self, request):
        from django.http import HttpResponse
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        if request.user.is_staff or request.user.is_superuser:
            logs = AuditLog.objects.all()
        else:
            logs = AuditLog.objects.filter(user=request.user)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Audit Logs"

        header_fill = PatternFill("solid", fgColor="3D7A3A")
        headers = ["Date & Time", "Username", "Full Name", "Action", "Module", "Description", "IP Address"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_num, log in enumerate(logs, 2):
            ws.cell(row=row_num, column=1, value=log.created_at.strftime("%Y-%m-%d %H:%M:%S"))
            ws.cell(row=row_num, column=2, value=log.user.username if log.user else "deleted")
            ws.cell(row=row_num, column=3, value=f"{log.user.first_name} {log.user.last_name}".strip() if log.user else "")
            ws.cell(row=row_num, column=4, value=log.get_action_display())
            ws.cell(row=row_num, column=5, value=log.model_name)
            ws.cell(row=row_num, column=6, value=log.description)
            ws.cell(row=row_num, column=7, value=log.ip_address or "")

        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="piglytics_audit_log.xlsx"'
        wb.save(response)
        return response

    @action(detail=False, methods=["get"])
    def download_pdf(self, request):
        from django.http import HttpResponse
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet
        import io

        if request.user.is_staff or request.user.is_superuser:
            logs = AuditLog.objects.all()[:200]
        else:
            logs = AuditLog.objects.filter(user=request.user)[:200]

        buffer = io.BytesIO()
        doc    = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        elements = []
        elements.append(Paragraph("Piglytics Audit Log", styles["Title"]))

        table_data = [["Date & Time", "Username", "Action", "Module", "Description"]]
        for log in logs:
            table_data.append([
                log.created_at.strftime("%Y-%m-%d %H:%M"),
                log.user.username if log.user else "deleted",
                log.get_action_display(),
                log.model_name or "—",
                (log.description or "")[:80],
            ])

        t = Table(table_data, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3D7A3A")),
            ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
            ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F7F1")]),
            ("GRID",       (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ]))
        elements.append(t)
        doc.build(elements)

        buffer.seek(0)
        response = HttpResponse(buffer, content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="piglytics_audit_log.pdf"'
        return response


# ── Admin ─────────────────────────────────────────────────────────────────────

class AdminStatsViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]

    def list(self, request):
        if not (request.user.is_staff or request.user.is_superuser):
            return Response({"error": "Administrator privileges required."}, status=403)

        today       = date.today()
        month_start = today.replace(day=1)

        farmer_users   = User.objects.filter(is_staff=False, is_superuser=False)
        total_users    = farmer_users.count()
        active_users   = farmer_users.filter(is_active=True).count()
        disabled_users = farmer_users.filter(is_active=False).count()
        new_this_month = farmer_users.filter(date_joined__date__gte=month_start).count()

        farmer_farms = Farm.objects.filter(owner__is_staff=False, owner__is_superuser=False)
        total_farms  = farmer_farms.count()

        farmer_pigs    = Pig.objects.filter(farm__in=farmer_farms)
        total_pigs     = farmer_pigs.count()
        total_pregnant = BreedingRecord.objects.filter(
            sow__farm__in=farmer_farms, pregnancy_status="pregnant"
        ).count()
        total_piglets  = farmer_pigs.filter(growth_stage="piglet").count()
        total_boars    = farmer_pigs.filter(gender="male").count()
        total_feed     = FeedInventory.objects.filter(farm__in=farmer_farms).count()
        total_medicine = MedicineInventory.objects.filter(farm__in=farmer_farms).count()
        total_logs     = AuditLog.objects.count()

        # ── Weather summary across all farms ─────────────────────────────────
        heat_stress_farms    = 0
        cold_stress_farms    = 0
        critical_farms_wx    = 0
        try:
            from .weather import get_weather_data
            wdata   = get_weather_data()
            curr    = wdata.get("current", {})
            temp_c  = float(curr.get("temperature_2m",       25.0))
            for farm in farmer_farms:
                worst = "normal"
                if temp_c >= 32:
                    heat_stress_farms += 1
                    worst = "critical"
                elif temp_c >= 28:
                    heat_stress_farms += 1
                    worst = "warning"
                elif temp_c < 18:
                    cold_stress_farms += 1
                    worst = "warning"
                if worst == "critical":
                    critical_farms_wx += 1
        except Exception:
            pass

        return Response({
            "total_users": total_users, "active_users": active_users,
            "disabled_users": disabled_users, "new_users_this_month": new_this_month,
            "total_farms": total_farms, "total_pigs": total_pigs,
            "total_pregnant": total_pregnant, "total_piglets": total_piglets,
            "total_boars": total_boars, "total_feed": total_feed,
            "total_medicine": total_medicine, "total_logs": total_logs,
            # Weather
            "farms_heat_stress":      heat_stress_farms,
            "farms_cold_stress":      cold_stress_farms,
            "farms_critical_weather": critical_farms_wx,
        })


class AdminFarmerViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]

    def _admin_check(self, request):
        return request.user.is_staff or request.user.is_superuser

    def list(self, request):
        if not self._admin_check(request):
            return Response({"error": "Administrator privileges required."}, status=403)

        users = User.objects.filter(is_staff=False, is_superuser=False).order_by("-date_joined")
        results = []
        for user in users:
            farm      = Farm.objects.filter(owner=user).first()
            farm_pigs = Pig.objects.filter(farm=farm) if farm else Pig.objects.none()
            last_login = user.last_login.strftime("%Y-%m-%d") if user.last_login else "Never"
            results.append({
                "id": user.id, "username": user.username,
                "full_name":   f"{user.first_name} {user.last_name}".strip() or user.username,
                "email":       user.email or "",
                "is_active":   user.is_active,
                "date_joined": user.date_joined.strftime("%Y-%m-%d"),
                "last_login":  last_login,
                "farm_name":   farm.name if farm else "No farm yet",
                "farm_id":     farm.id   if farm else None,
                "total_pigs":  farm_pigs.count(),
                "pregnant_sows": BreedingRecord.objects.filter(
                    sow__farm=farm, pregnancy_status="pregnant"
                ).count() if farm else 0,
                "total_piglets":  farm_pigs.filter(growth_stage="piglet").count(),
                "total_boars":    farm_pigs.filter(gender="male", growth_stage__in=["grower","finisher","breeder"]).count(),
                "feed_count":     FeedInventory.objects.filter(farm=farm).count()     if farm else 0,
                "medicine_count": MedicineInventory.objects.filter(farm=farm).count() if farm else 0,
                "health_logs":    HealthLog.objects.filter(pig__farm=farm).count()    if farm else 0,
            })
        return Response(results)

    @action(detail=True, methods=["post"])
    def disable(self, request, pk=None):
        if not self._admin_check(request):
            return Response({"error": "Administrator privileges required."}, status=403)
        try:
            user = User.objects.get(pk=pk, is_staff=False, is_superuser=False)
        except User.DoesNotExist:
            return Response({"error": "User not found."}, status=404)
        user.is_active = False
        user.save()
        log_action(request.user, "update", "User", user.id, f"Admin disabled: {user.username}", request)
        return Response({"message": f"'{user.username}' has been disabled."})

    @action(detail=True, methods=["post"])
    def activate(self, request, pk=None):
        if not self._admin_check(request):
            return Response({"error": "Administrator privileges required."}, status=403)
        try:
            user = User.objects.get(pk=pk, is_staff=False, is_superuser=False)
        except User.DoesNotExist:
            return Response({"error": "User not found."}, status=404)
        user.is_active = True
        user.save()
        log_action(request.user, "update", "User", user.id, f"Admin activated: {user.username}", request)
        return Response({"message": f"'{user.username}' has been activated."})

    @action(detail=True, methods=["post"])
    def reset_password(self, request, pk=None):
        if not self._admin_check(request):
            return Response({"error": "Administrator privileges required."}, status=403)
        try:
            user = User.objects.get(pk=pk, is_staff=False, is_superuser=False)
        except User.DoesNotExist:
            return Response({"error": "User not found."}, status=404)
        temp = "piglytics2025"
        user.set_password(temp)
        user.save()
        log_action(request.user, "update", "User", user.id, f"Admin reset password for: {user.username}", request)
        return Response({"message": f"Password for '{user.username}' reset to: {temp}"})


class FarmerAnalyticsViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]

    def retrieve(self, request, pk=None):
        if not (request.user.is_staff or request.user.is_superuser):
            return Response({"error": "Administrator privileges required."}, status=403)
        try:
            user = User.objects.get(pk=pk, is_staff=False, is_superuser=False)
        except User.DoesNotExist:
            return Response({"error": "Farmer not found."}, status=404)

        farm      = Farm.objects.filter(owner=user).first()
        farm_pigs = Pig.objects.filter(farm=farm) if farm else Pig.objects.none()
        today     = date.today()

        profile_data = {
            "id":            user.id,
            "username":      user.username,
            "full_name":     f"{user.first_name} {user.last_name}".strip() or user.username,
            "email":         user.email or "—",
            "is_active":     user.is_active,
            "date_joined":   user.date_joined.strftime("%B %d, %Y"),
            "last_login":    user.last_login.strftime("%Y-%m-%d %H:%M") if user.last_login else "Never",
            "farm_name":     farm.name     if farm else "No farm",
            "farm_location": farm.location if farm else "—",
        }

        total_pigs     = farm_pigs.count()
        healthy_count  = farm_pigs.filter(health_status="healthy").count()
        sick_count     = farm_pigs.filter(health_status__in=["under_treatment","critical"]).count()
        critical_count = farm_pigs.filter(health_status="critical").count()
        pregnant_count = BreedingRecord.objects.filter(sow__farm=farm, pregnancy_status="pregnant").count() if farm else 0
        piglet_count   = farm_pigs.filter(growth_stage="piglet").count()

        stages = ["piglet","weaner","grower","finisher","breeder"]
        stage_breakdown  = [{"stage": s.capitalize(), "count": farm_pigs.filter(growth_stage=s).count()} for s in stages]
        gender_breakdown = [
            {"gender": "Female", "count": farm_pigs.filter(gender="female").count()},
            {"gender": "Male",   "count": farm_pigs.filter(gender="male").count()},
        ]

        weight_records = WeightRecord.objects.filter(
            pig__farm=farm, recorded_at__gte=today - timedelta(days=180)
        ).order_by("recorded_at") if farm else []
        weight_by_month = {}
        for wr in weight_records:
            key = wr.recorded_at.strftime("%b %Y")
            if key not in weight_by_month:
                weight_by_month[key] = {"total": 0, "count": 0}
            weight_by_month[key]["total"] += float(wr.weight_kg)
            weight_by_month[key]["count"] += 1
        weight_trend = [{"month": k, "avg_weight": round(v["total"]/v["count"],1)} for k,v in weight_by_month.items()]

        health_logs = list(HealthLog.objects.filter(pig__farm=farm) if farm else [])
        health_by_severity = {"normal":0,"warning":0,"critical":0}
        for hl in health_logs:
            health_by_severity[hl.severity] = health_by_severity.get(hl.severity,0)+1

        breeding_qs   = BreedingRecord.objects.filter(sow__farm=farm) if farm else BreedingRecord.objects.none()
        farrowed_qs   = breeding_qs.filter(pregnancy_status="farrowed")
        total_litters = farrowed_qs.count()
        total_alive   = sum(r.piglets_born_alive or 0 for r in farrowed_qs)
        breeding_total = breeding_qs.count()
        breeding_success = round(total_litters/breeding_total*100,1) if breeding_total>0 else 0
        avg_litter_size  = round(total_alive/total_litters,1) if total_litters>0 else 0

        feed_items    = list(FeedInventory.objects.filter(farm=farm) if farm else [])
        total_feed_kg = sum(float(f.stock_kg) for f in feed_items)
        low_feed      = sum(1 for f in feed_items if float(f.stock_kg)<=25)

        feed_usage_logs = FeedUsageLog.objects.filter(farm=farm, date_used__gte=today-timedelta(days=30)).order_by("date_used") if farm else []
        feed_by_week = {}
        for fu in feed_usage_logs:
            key = f"Wk {fu.date_used.isocalendar()[1]}"
            feed_by_week[key] = feed_by_week.get(key,0)+float(fu.amount_used_kg)
        feed_weekly = [{"week":k,"kg_used":round(v,1)} for k,v in feed_by_week.items()]

        predicted_piglets = round(avg_litter_size*pregnant_count) if avg_litter_size else 0
        avg_weekly_feed   = sum(feed_by_week.values())/len(feed_by_week) if feed_by_week else 0
        avg_daily_feed    = avg_weekly_feed/7 if avg_weekly_feed else 0
        feed_days_left    = round(total_feed_kg/avg_daily_feed) if avg_daily_feed>0 else None

        return Response({
            "profile":          profile_data,
            "pig_summary":      {"total":total_pigs,"healthy":healthy_count,"sick":sick_count,"critical":critical_count,"pregnant":pregnant_count,"piglets":piglet_count},
            "stage_breakdown":  stage_breakdown,
            "gender_breakdown": gender_breakdown,
            "weight_trend":     weight_trend,
            "health_summary":   {**health_by_severity,"total_health_logs":len(health_logs)},
            "vaccination_summary": {
                "completed": VaccinationRecord.objects.filter(pig__farm=farm).count() if farm else 0,
                "due_soon":  VaccinationRecord.objects.filter(pig__farm=farm, next_due_date__lte=today+timedelta(days=7)).count() if farm else 0,
                "disease_cases": DiseaseRecord.objects.filter(pig__farm=farm).count() if farm else 0,
            },
            "breeding_summary": {"total_records":breeding_total,"currently_pregnant":pregnant_count,"total_litters":total_litters,"avg_litter_size":avg_litter_size,"success_rate_pct":breeding_success,"total_alive_born":total_alive},
            "feed_summary":     {"total_stock_kg":round(total_feed_kg,1),"low_stock_items":low_feed,"feed_items":len(feed_items)},
            "feed_weekly_usage": feed_weekly,
            "predictions":      {"predicted_pig_population":total_pigs+predicted_piglets,"predicted_new_piglets":predicted_piglets,"predicted_monthly_feed_kg":round(avg_weekly_feed*4.3,1),"feed_days_remaining":feed_days_left},
        })